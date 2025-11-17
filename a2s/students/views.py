from django.contrib.auth import authenticate, login, logout
from django.views.decorators.http import require_POST
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from supabase import create_client
from datetime import datetime, date, timedelta
import json
import os
from io import BytesIO

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import StudentProfile, Enrollment, Grade,Schedule
from students.models import Curriculum
from authentication.models import User


# Supabase Configuration & Constants

SUPABASE_URL = settings.SUPABASE_URL
SUPABASE_SERVICE_KEY = settings.SUPABASE_SERVICE_KEY
DEFAULT_PROFILE_URL = f"{SUPABASE_URL}/storage/v1/object/public/ProfilePicture/avatar.png"
supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)



# Base View
@login_required(login_url="Login")
def student_base(request):
    return render(request, "students/student_base.html")


# ----------------------------------------------------------------------
# Student Dashboard & Profile Views
# ----------------------------------------------------------------------
@login_required(login_url="Login")
def student_dashboard(request):
    student_user = request.user
    profile, _ = StudentProfile.objects.get_or_create(user=student_user)
    student_program = profile.program or "Undeclared"

    curriculum_data, calendar_data, current_courses = [], [], []

   
    total_units = 0
    completed_units = 0
    completion_percentage = 0.0
    credits_remaining = 0

   
    grades = profile.grades.all()
    gpa_total_points = 0
    gpa_total_units = 0

    print(f"Calculating GPA for student: {student_user.get_full_name()}")

    for grade in grades:
       
        if grade.status and grade.status.upper() != "CURRENT":
            continue

        units = grade.units or 3

        
        midterm_score = float(grade.midterm or 0)
        final_score = float(grade.final) if grade.final is not None else midterm_score
        course_grade = (midterm_score + final_score) / 2

        gpa_total_points += course_grade * units
        gpa_total_units += units

    current_gpa = round(gpa_total_points / gpa_total_units, 2) if gpa_total_units > 0 else 0.0
    print(f"Calculated GPA: {current_gpa}")

    try:
        if student_program != "Undeclared":
            curriculum_obj = Curriculum.objects.get(program=student_program)
            curriculum_data = curriculum_obj.data.get("curriculum", [])

            for year in curriculum_data:
                for term in year.get("terms", []):
                    for subject in term.get("subjects", []):
                        units = subject.get('units', 3)
                        total_units += units
                        status = subject.get('final_grade', 'RECOMMENDED')
                        if status.upper() == 'PASSED':
                            completed_units += units

            if total_units > 0:
                completion_percentage = round((completed_units / total_units * 100), 1)
            credits_remaining = total_units - completed_units

           
            for year in curriculum_data:
                for term in year.get("terms", []):
                    for subj in term.get("subjects", []):
                        if subj.get("final_grade", "").upper() == "CURRENT":
                            current_courses.append(subj)
    except Curriculum.DoesNotExist:
        pass


    import os, json
    from django.conf import settings
    calendar_path = os.path.join(settings.BASE_DIR, "static", "data", "academic_calendar.json")
    try:
        with open(calendar_path, "r", encoding="utf-8") as f:
            calendar_data = json.load(f)
    except Exception as e:
        print("Error loading academic calendar:", e)

    from datetime import date, datetime
    today = date.today()
    current_semester = "Unknown Semester"

    for sem in calendar_data:
        start_event = next((e for e in sem["events"] if "Start of Classes" in e["name"]), None)
        end_event = next((e for e in sem["events"] if "End of Classes" in e["name"]), None)
        if start_event and end_event:
            start_date = datetime.strptime(start_event["date"], "%Y-%m-%d").date()
            end_date = datetime.strptime(end_event["date"], "%Y-%m-%d").date()
            if start_date <= today <= end_date:
                current_semester = sem["semester"]
                break

    context = {
        "student": student_user,
        "student_program": student_program,
        "curriculum_json": json.dumps(curriculum_data),
        "calendar_json": json.dumps(calendar_data),
        "current_courses": current_courses,
        "active_courses_count": len(current_courses),
        "gpa": current_gpa,
        "credits_completed": profile.credits_completed or 0,
        "credits_required": profile.credits_required or 0,
        "year_level": profile.year_level,
        "academic_standing": profile.academic_standing,
        "current_semester": current_semester,
       
        "total_units": total_units,
        "completed_units": completed_units,
        "completion_percentage": completion_percentage,
        "credits_remaining": credits_remaining,
        "current_gpa": current_gpa,
        "profile": profile,
    }

    return render(request, "students/StudentDashboard.html", context)

@login_required(login_url="Login")
def student_profile(request):
    user = request.user
    profile, _ = StudentProfile.objects.get_or_create(user=user)

    # --- GPA calculation ---
    grades = profile.grades.all()
    gpa_total_points = 0
    gpa_total_units = 0
    for grade in grades:
        if grade.status and grade.status.upper() != "CURRENT":
            continue
        units = grade.units or 3
        midterm_score = float(grade.midterm or 0)
        final_score = float(grade.final) if grade.final is not None else midterm_score
        course_grade = (midterm_score + final_score) / 2
        gpa_total_points += course_grade * units
        gpa_total_units += units
    current_gpa = round(gpa_total_points / gpa_total_units, 2) if gpa_total_units > 0 else 0.0

    # --- Credits calculation ---
    total_units = 0
    completed_units = 0
    try:
        if profile.program:
            curriculum_obj = Curriculum.objects.get(program=profile.program)
            curriculum_data = curriculum_obj.data.get("curriculum", [])
            for year in curriculum_data:
                for term in year.get("terms", []):
                    for subject in term.get("subjects", []):
                        units = subject.get('units', 3)
                        total_units += units
                        if subject.get("final_grade","").upper() == "PASSED":
                            completed_units += units
        credits_remaining = total_units - completed_units
        progress_percentage = round((completed_units / total_units * 100), 1) if total_units else 0
    except Curriculum.DoesNotExist:
        total_units = completed_units = credits_remaining = progress_percentage = 0

    def format_date(dt):
        return dt.strftime("%B %d, %Y") if dt else ""

    context = {
        "first_name": user.first_name,
        "last_name": user.last_name,
        "email": user.email,
        "phone": profile.phone or "",
        "address": profile.address or "",
        "dob": profile.dob or "",
        "dob_display": format_date(profile.dob),
        "bio": profile.bio or "",
        "student_id": user.id_number,
        "major": profile.program or "",
        "academic_year": profile.year_level or 1,
        "expected_graduation": profile.expected_graduation or "",
        "expected_graduation_display": format_date(profile.expected_graduation),

        "current_gpa": current_gpa,  # use current_gpa
        "credits_completed": completed_units,
        "credits_required": total_units,
        "academic_standing": profile.academic_standing or "Good Standing",
        "progress_percentage": progress_percentage,

        "achievements": profile.achievements.all(),
        "profile_picture_url": profile.profile_picture or "https://qimrryerxdzfewbkoqyq.supabase.co/storage/v1/object/public/ProfilePicture/avatar.png",
    }

    return render(request, "students/StudentProfile.html", context)


@csrf_exempt 
@login_required(login_url="Login")
def update_student_profile(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            user = request.user
            student = StudentProfile.objects.get(user=user)

            user.first_name = data.get("first_name", user.first_name)
            user.last_name = data.get("last_name", user.last_name)
            user.email = data.get("email", user.email)
            user.save()

            student.phone = data.get("phone", student.phone)
            student.address = data.get("address", student.address)
            student.bio = data.get("bio", student.bio)
            student.program = data.get("major", student.program)
            student.year_level = data.get("academic_year", student.year_level)
            student.gpa = data.get("gpa", student.gpa)
            student.credits_completed = data.get("credits_completed", student.credits_completed)
            student.academic_standing = data.get("academic_standing", student.academic_standing)
            student.save()

            return JsonResponse({"status": "success"})
        except Exception as e:
            print("❌ ERROR updating profile:", e)
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)


@csrf_exempt
@login_required
def upload_student_profile_picture(request):
    if request.method != 'POST' or 'file' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

    file = request.FILES['file']
    user = request.user

    try:
        profile, _ = StudentProfile.objects.get_or_create(user=user)
        file_path = f"{user.id}/{file.name}"
        file_bytes = file.read()

        upload_result = supabase.storage.from_('ProfilePicture').upload(file_path, file_bytes)

        if hasattr(upload_result, 'error') and upload_result.error is not None:
            raise Exception(f"Supabase upload error: {upload_result.error}")

        url_result = supabase.storage.from_('ProfilePicture').get_public_url(file_path)
        public_url = getattr(url_result, 'public_url', None) or getattr(url_result, 'url', None)

        if not public_url:
            public_url = f"{SUPABASE_URL}/storage/v1/object/public/ProfilePicture/{file_path}"

        profile.profile_picture = public_url
        profile.save(update_fields=['profile_picture'])

        return JsonResponse({'status': 'success', 'url': public_url})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def reset_student_profile_picture(request):
    profile, _ = StudentProfile.objects.get_or_create(user=request.user)
    default_url = "https://qimrryerxdzfewbkoqyq.supabase.co/storage/v1/object/public/ProfilePicture/avatar.png"
    profile.profile_picture = default_url
    profile.save(update_fields=['profile_picture'])
    return JsonResponse({'status': 'success', 'url': default_url})


# ----------------------------------------------------------------------
# Schedule & Course Enrollment Views
# ----------------------------------------------------------------------


@login_required(login_url="Login")
def student_schedule(request):
    days = ['M', 'T', 'W', 'TH', 'F', 'SAT']

    start_time = datetime.strptime("07:00 AM", "%I:%M %p")
    end_time = datetime.strptime("09:00 PM", "%I:%M %p")
    hours = []
    current = start_time
    while current <= end_time:
        hours.append(current.strftime("%I:%M %p"))
        current += timedelta(minutes=30)

    raw_scheds = Schedule.objects.filter(student__user=request.user)

    schedule_grid = {day: {} for day in days}

    for s in raw_scheds:
        day = s.day.strip().upper()
        start_dt = datetime.combine(datetime.today(), s.time_start)
        end_dt = datetime.combine(datetime.today(), s.time_end)

        blocks = int((end_dt - start_dt).total_seconds() // (30 * 60))
        current_dt = start_dt
        for i in range(blocks):
            time_str = current_dt.strftime("%I:%M %p")
            if i == 0:
                schedule_grid[day][time_str] = {
                    'code': s.code,
                    'section': s.section,
                    'room': s.room,
                    'blocks': blocks,
                }
            else:
                schedule_grid[day][time_str] = {'occupied': True}
            current_dt += timedelta(minutes=30)

    context = {
        'days': days,
        'hours': hours,
        'schedule_grid': schedule_grid,
    }
    return render(request, 'students/StudentSchedule.html', context)


@login_required(login_url="Login")
def student_curriculum(request):
    profile, _ = StudentProfile.objects.get_or_create(user=request.user)
    enrollments = Enrollment.objects.filter(student=profile).select_related("course")
    return render(request, "students/StudentCurriculum.html", {"enrollments": enrollments})


@login_required(login_url='Login')
def student_courses(request):
    
    student_user = request.user
    profile, _ = StudentProfile.objects.get_or_create(user=student_user)
    student_program = profile.program if profile.program and profile.program != "Undeclared" else None

    calendar_path = os.path.join(settings.BASE_DIR, 'a2s', 'static', 'data', 'academic_calendar.json')
    current_semester = None
    current_phase = None

    try:
        with open(calendar_path, 'r', encoding='utf-8') as f:
            academic_calendar = json.load(f)

        now = datetime.now().date()

        for sem in academic_calendar:
            events = sem.get("events", [])
            start = next((datetime.strptime(e["date"], "%Y-%m-%d").date() for e in events if "Start of Classes" in e["name"]), None)
            end = next((datetime.strptime(e["date"], "%Y-%m-%d").date() for e in events if "End of Classes" in e["name"]), None)

            if start and end and start <= now <= end:
                current_semester = sem.get("semester")

                prelim_end = next((datetime.strptime(e["date"], "%Y-%m-%d").date() for e in events if "Prelim Examinations" in e["name"]), None)
                midterm_end = next((datetime.strptime(e["date"], "%Y-%m-%d").date() for e in events if "Midterm Examinations" in e["name"]), None)
                finals_end = end

                if start and prelim_end and start <= now <= prelim_end:
                    current_phase = "Prelim"
                elif prelim_end and midterm_end and prelim_end < now <= midterm_end:
                    current_phase = "Midterm"
                elif midterm_end and end and midterm_end < now <= end:
                    current_phase = "Finals"
                elif end and now > end:
                    current_phase = "Completed"
                else:
                    current_phase = "Before Semester"

                
    except Exception as e:
        print("Error reading academic calendar:", e)

    curriculum_data = []
    active_courses_count = 0
    if student_program:
        try:
            curriculum_obj = Curriculum.objects.get(program=student_program)
            curriculum_data = curriculum_obj.data.get("curriculum", [])
            for year in curriculum_data:
                for term in year.get("terms", []):
                    for sub in term.get("subjects", []):
                        if sub.get("final_grade") == "CURRENT":
                            active_courses_count += 1
        except Curriculum.DoesNotExist:
            pass

    context = {
        "student": student_user,
        "student_program": student_program or "No program assigned",
        "curriculum_json": json.dumps(curriculum_data),
        "current_semester": current_semester or "No Active Semester",
        "current_phase": current_phase or "N/A",
        "current_year": datetime.now().year,
        "active_courses": active_courses_count,
    }

    return render(request, "students/StudentCourses.html", context)


# ----------------------------------------------------------------------
# Degree Audit & Grades Views
# ----------------------------------------------------------------------
@login_required(login_url='Login')
def student_degree_audit(request):
    student_user = request.user
    profile, _ = StudentProfile.objects.get_or_create(user=student_user)
    student_program = profile.program if profile.program and profile.program != "Undeclared" else None

    curriculum_data = []
    all_courses = []
    completed_courses = []
    in_progress_courses = []
    not_started_courses = []
    
    gen_ed_courses = []
    major_core_courses = []
    elective_courses = []
    minor_courses = []
    
    total_units = 0
    completed_units = 0
    in_progress_units = 0
    
    if student_program:
        try:
            curriculum_obj = Curriculum.objects.get(program=student_program)
            curriculum_data = curriculum_obj.data.get("curriculum", [])
            
            for year in curriculum_data:
                for term in year.get("terms", []):
                    for subject in term.get("subjects", []):
                        course_info = {
                            'code': subject.get('subject_code', 'N/A'),
                            'title': subject.get('description', 'N/A'),
                            'units': subject.get('units', 3),
                            'semester': f"{year.get('year', '')} - {term.get('term', '')}",
                            'status': subject.get('final_grade', 'RECOMMENDED'),
                            'school_year': subject.get('school_year', ''),
                            'year': year.get('year', '')
                        }
                        
                        all_courses.append(course_info)
                        total_units += course_info['units']
                        
                        if course_info['status'] == 'PASSED':
                            completed_courses.append(course_info)
                            completed_units += course_info['units']
                        elif course_info['status'] == 'CURRENT':
                            in_progress_courses.append(course_info)
                            in_progress_units += course_info['units']
                        else:
                            not_started_courses.append(course_info)
                        
                        code = course_info['code'].upper()
                        
                        if any(prefix in code for prefix in ['GE', 'ENGL', 'MATH', 'PHILO', 'HUM', 'SOCSCI', 'NATSCI', 'HIST', 'PSYCH', 'PE', 'NSTP']):
                            gen_ed_courses.append(course_info)
                        elif 'ELEC' in code or 'ELCE' in code or 'FREEEL' in code:
                            elective_courses.append(course_info)
                        elif 'IT' in code or 'CSIT' in code or 'CS' in code:
                            major_core_courses.append(course_info)
                        else:
                            minor_courses.append(course_info)
                            
        except Curriculum.DoesNotExist:
            pass
    
    completion_percentage = (completed_units / total_units * 100) if total_units > 0 else 0
    credits_remaining = total_units - completed_units
    
    current_gpa = float(profile.gpa) if profile.gpa else 0.0
    
    achievements = profile.achievements.all()
    
    recommended_courses = []
    
    current_year_level = profile.year_level
    if curriculum_data and current_year_level <= len(curriculum_data):
        current_year_data = curriculum_data[current_year_level - 1] if current_year_level > 0 else None
        
        if current_year_data:
            for term in current_year_data.get('terms', []):
                for subject in term.get('subjects', []):
                    if subject.get('final_grade', '').upper() == 'RECOMMENDED':
                        recommended_courses.append({
                            'code': subject.get('subject_code', 'N/A'),
                            'title': subject.get('description', 'N/A'),
                            'units': subject.get('units', 3),
                            'semester': term.get('term', 'N/A')
                        })
    
    recommended_courses = recommended_courses[:6]
    
    calendar_path = os.path.join(settings.BASE_DIR, 'a2s', 'static', 'data', 'academic_calendar.json')
    current_semester = "Unknown Semester"
    try:
        with open(calendar_path, 'r', encoding='utf-8') as f:
            calendar_data = json.load(f)
            today = datetime.now().date()
            
            for sem in calendar_data:
                start_event = next((e for e in sem["events"] if "Start of Classes" in e["name"]), None)
                end_event = next((e for e in sem["events"] if "End of Classes" in e["name"]), None)
                
                if start_event and end_event:
                    start_date = datetime.strptime(start_event["date"], "%Y-%m-%d").date()
                    end_date = datetime.strptime(end_event["date"], "%Y-%m-%d").date()
                    
                    if start_date <= today <= end_date:
                        current_semester = sem["semester"]
                        break
    except Exception as e:
        print("Error loading academic calendar:", e)
    
    context = {
        'student': student_user,
        'profile': profile,
        'student_program': student_program or "No Program Assigned",
        'current_semester': current_semester,
        
        'total_units': total_units,
        'completed_units': completed_units,
        'in_progress_units': in_progress_units,
        'credits_remaining': credits_remaining,
        'completion_percentage': round(completion_percentage, 1),
        'current_gpa': current_gpa,
        
        'gen_ed_courses': gen_ed_courses,
        'major_core_courses': major_core_courses,
        'elective_courses': elective_courses,
        'minor_courses': minor_courses,
        
        'completed_courses': completed_courses,
        'in_progress_courses': in_progress_courses,
        'not_started_courses': not_started_courses,
        
        'achievements': achievements,
        'recommended_courses': recommended_courses,
        
        'all_courses_json': json.dumps([{
            'code': c['code'],
            'title': c['title'],
            'units': c['units']
        } for c in all_courses]),
    }
    
    return render(request, 'students/StudentDegreeAudit.html', context)


@login_required(login_url="Login")
def student_grades(request):
    profile, _ = StudentProfile.objects.get_or_create(user=request.user)
    selected_year = request.GET.get("school_year")
    selected_sem = request.GET.get("semester")

    grades = Grade.objects.none()
    years = Grade.objects.filter(student=profile).values_list("school_year", flat=True).distinct()
    semesters = Grade.objects.filter(student=profile).values_list("semester", flat=True).distinct()

    if selected_year and selected_sem:
        grades = Grade.objects.filter(student=profile, school_year=selected_year, semester=selected_sem)

    return render(request, "students/StudentGrades.html", {
        "grades": grades,
        "years": years,
        "semesters": semesters,
        "selected_year": selected_year,
        "selected_semester": selected_sem,
    })


# ----------------------------------------------------------------------
# Utility and API Views (Progress, Curriculum, JSON)
# ----------------------------------------------------------------------
@login_required(login_url="Login")
def get_curriculum_json(request, program):
    try:
        curriculum = Curriculum.objects.get(program__iexact=program)
        return JsonResponse(curriculum.data, safe=False)
    except Curriculum.DoesNotExist:
        return JsonResponse({'error': f'No curriculum found for {program}'}, status=404)
    

@login_required(login_url="Login")
def student_schedule_json(request, student_id):
    student_profile = StudentProfile.objects.get(id=student_id)
    enrollments = Enrollment.objects.filter(student=student_profile)

    schedule = []
    for e in enrollments:
        schedule.append({
            "code": e.course.course_code,
            "section": getattr(e.course, 'section', 'N/A'),
            "room": getattr(e, 'room', 'N/A'),
            "time": f"{e.schedule_day} {e.start_time.strftime('%I:%M%p')}-{e.end_time.strftime('%I:%M%p')}"
        })

    return JsonResponse({"student_id": student_id, "schedule": schedule})


@login_required(login_url="Login")
def calculate_semester_progress(calendar_data):
    today = datetime.now().date()
    for semester in calendar_data:
        events = semester.get("events", [])
        start = next((datetime.strptime(e["date"], "%Y-%m-%d").date() for e in events if e["type"] == "Class Start"), None)
        end = next((datetime.strptime(e["date"], "%Y-%m-%d").date() for e in events if e["type"] == "Class End"), None)
        if start and end and start <= today <= end:
            total = (end - start).days
            elapsed = (today - start).days
            return max(0, min(int((elapsed / total) * 100), 100))
    return 0


@login_required(login_url="Login")
def calculate_subject_progress(semester_name, subject_status):
    if subject_status == "passed":
        return 100
    elif subject_status == "recommended":
        return 0
    elif subject_status == "current":
        with open("path/to/academic_calendar.json") as f:
            data = json.load(f)

        semester = next((s for s in data if s["semester"] == semester_name), None)
        if not semester:
            return 0

        start_event = next((e for e in semester["events"] if e["type"] == "Class Start"), None)
        end_event = next((e for e in semester["events"] if e["type"] == "Class End"), None)

        if not start_event or not end_event:
            return 0

        start_date = datetime.strptime(start_event["date"], "%Y-%m-%d").date()
        end_date = datetime.strptime(end_event["date"], "%Y-%m-%d").date()
        today = date.today()

        if today < start_date:
            return 0
        elif today > end_date:
            return 100

        total_days = (end_date - start_date).days
        elapsed_days = (today - start_date).days
        progress = (elapsed_days / total_days) * 100
        return round(progress, 1)


@login_required(login_url="Login")
def get_current_phase(academic_calendar):
    today = date.today()

    for sem in academic_calendar:
        events = sem['events']
        for e in events:
            if e['type'] == 'Exam':
                event_date = datetime.strptime(e['date'], "%Y-%m-%d").date()
                if today <= event_date:
                    return {
                        "semester": sem['semester'],
                        "phase": e['name'].split()[0] 
                    }
    return {"semester": None, "phase": "Completed"}

@login_required
def get_notifications(request):
    student = request.user.studentprofile
    notifications = student.notifications.order_by('-date_created')
    data = [
        {
            "id": n.id,
            "title": n.title,
            "message": n.message,
            "type": n.type,
            "read": n.read,
            "date": n.date_created.strftime("%Y-%m-%d %H:%M"),
        }
        for n in notifications
    ]
    unread_count = student.notifications.filter(read=False).count()
    return JsonResponse({"notifications": data, "unread_count": unread_count})


@login_required
@require_POST
def mark_notification_read(request):
    student = request.user.studentprofile
    notif_id = request.POST.get("id")
    try:
        notif = Notification.objects.get(id=notif_id, student=student)
        notif.read = True
        notif.save()
        return JsonResponse({"success": True})
    except Notification.DoesNotExist:
        return JsonResponse({"success": False, "error": "Notification not found"})

# ----------------------------------------------------------------------
# Export Degree Audit to PDF
# ----------------------------------------------------------------------
@login_required(login_url='Login')
def export_degree_audit_pdf(request):
    """Generate a PDF report of the student's degree audit"""
    student_user = request.user
    profile, _ = StudentProfile.objects.get_or_create(user=student_user)
    student_program = profile.program if profile.program and profile.program != "Undeclared" else "No Program"
    
    # Fetch curriculum data
    curriculum_data = []
    all_courses = []
    completed_courses = []
    in_progress_courses = []
    not_started_courses = []
    
    total_units = 0
    completed_units = 0
    
    if profile.program and profile.program != "Undeclared":
        try:
            curriculum_obj = Curriculum.objects.get(program=profile.program)
            curriculum_data = curriculum_obj.data.get("curriculum", [])
            
            for year in curriculum_data:
                for term in year.get("terms", []):
                    for subject in term.get("subjects", []):
                        course_info = {
                            'code': subject.get('subject_code', 'N/A'),
                            'title': subject.get('description', 'N/A'),
                            'units': subject.get('units', 3),
                            'status': subject.get('final_grade', 'RECOMMENDED'),
                        }
                        
                        all_courses.append(course_info)
                        total_units += course_info['units']
                        
                        if course_info['status'] == 'PASSED':
                            completed_courses.append(course_info)
                            completed_units += course_info['units']
                        elif course_info['status'] == 'CURRENT':
                            in_progress_courses.append(course_info)
                        else:
                            not_started_courses.append(course_info)
        except Curriculum.DoesNotExist:
            pass
    
    completion_percentage = (completed_units / total_units * 100) if total_units > 0 else 0
    credits_remaining = total_units - completed_units
    current_gpa = float(profile.gpa) if profile.gpa else 0.0
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4f46e5'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    normal_style = styles['Normal']
    
    # Header
    elements.append(Paragraph("A2S Academic Advising System", title_style))
    elements.append(Paragraph("Degree Audit Report", heading_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Student Information
    student_info_data = [
        ['Student Name:', f"{student_user.first_name} {student_user.last_name}"],
        ['Student ID:', student_user.id_number or 'N/A'],
        ['Program:', student_program],
        ['Year Level:', f"Year {profile.year_level}" if profile.year_level else 'N/A'],
        ['Generated On:', datetime.now().strftime('%B %d, %Y at %I:%M %p')],
    ]
    
    student_info_table = Table(student_info_data, colWidths=[2*inch, 4*inch])
    student_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
    ]))
    
    elements.append(student_info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Progress Summary
    elements.append(Paragraph("Progress Summary", heading_style))
    
    progress_data = [
        ['Metric', 'Value'],
        ['Completion Percentage', f"{completion_percentage:.1f}%"],
        ['Credits Completed', f"{completed_units} / {total_units}"],
        ['Credits Remaining', str(credits_remaining)],
        ['Current GPA', f"{current_gpa:.2f}"],
        ['Expected Graduation', profile.expected_graduation.strftime('%B %Y') if profile.expected_graduation else 'TBD'],
    ]
    
    progress_table = Table(progress_data, colWidths=[3*inch, 3*inch])
    progress_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(progress_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Completed Courses
    if completed_courses:
        elements.append(Paragraph("Completed Courses", heading_style))
        
        completed_data = [['Course Code', 'Course Title', 'Units']]
        for course in completed_courses:
            completed_data.append([
                course['code'],
                course['title'][:50] + '...' if len(course['title']) > 50 else course['title'],
                str(course['units'])
            ])
        
        completed_table = Table(completed_data, colWidths=[1.5*inch, 3.5*inch, 1*inch])
        completed_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22c55e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf4')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(completed_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # In Progress Courses
    if in_progress_courses:
        elements.append(Paragraph("Courses In Progress", heading_style))
        
        progress_data = [['Course Code', 'Course Title', 'Units']]
        for course in in_progress_courses:
            progress_data.append([
                course['code'],
                course['title'][:50] + '...' if len(course['title']) > 50 else course['title'],
                str(course['units'])
            ])
        
        progress_table = Table(progress_data, colWidths=[1.5*inch, 3.5*inch, 1*inch])
        progress_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#facc15')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fefce8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(progress_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Remaining Courses
    if not_started_courses:
        elements.append(PageBreak())
        elements.append(Paragraph("Remaining Courses", heading_style))
        
        remaining_data = [['Course Code', 'Course Title', 'Units']]
        for course in not_started_courses:
            remaining_data.append([
                course['code'],
                course['title'][:50] + '...' if len(course['title']) > 50 else course['title'],
                str(course['units'])
            ])
        
        remaining_table = Table(remaining_data, colWidths=[1.5*inch, 3.5*inch, 1*inch])
        remaining_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(remaining_table)
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_text = "This is an official document generated by the A2S Academic Advising System."
    elements.append(Paragraph(footer_text, ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#6b7280'),
        alignment=TA_CENTER
    )))
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF from buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="degree_audit_{student_user.id_number}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response


# ----------------------------------------------------------------------
# Export Degree Audit to Excel
# ----------------------------------------------------------------------
@login_required(login_url='Login')
def export_degree_audit_excel(request):
    """Generate an Excel report of the student's degree audit"""
    student_user = request.user
    profile, _ = StudentProfile.objects.get_or_create(user=student_user)
    student_program = profile.program if profile.program and profile.program != "Undeclared" else "No Program"
    
    # Fetch curriculum data
    curriculum_data = []
    all_courses = []
    completed_courses = []
    in_progress_courses = []
    not_started_courses = []
    
    total_units = 0
    completed_units = 0
    
    if profile.program and profile.program != "Undeclared":
        try:
            curriculum_obj = Curriculum.objects.get(program=profile.program)
            curriculum_data = curriculum_obj.data.get("curriculum", [])
            
            for year in curriculum_data:
                for term in year.get("terms", []):
                    for subject in term.get("subjects", []):
                        course_info = {
                            'code': subject.get('subject_code', 'N/A'),
                            'title': subject.get('description', 'N/A'),
                            'units': subject.get('units', 3),
                            'status': subject.get('final_grade', 'RECOMMENDED'),
                        }
                        
                        all_courses.append(course_info)
                        total_units += course_info['units']
                        
                        if course_info['status'] == 'PASSED':
                            completed_courses.append(course_info)
                            completed_units += course_info['units']
                        elif course_info['status'] == 'CURRENT':
                            in_progress_courses.append(course_info)
                        else:
                            not_started_courses.append(course_info)
        except Curriculum.DoesNotExist:
            pass
    
    completion_percentage = (completed_units / total_units * 100) if total_units > 0 else 0
    credits_remaining = total_units - completed_units
    current_gpa = float(profile.gpa) if profile.gpa else 0.0
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Degree Audit"
    
    # Define styles
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    success_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    warning_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    neutral_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'A2S Academic Advising System - Degree Audit Report'
    ws['A1'].font = Font(bold=True, size=16, color='4F46E5')
    ws['A1'].alignment = center_align
    
    # Student Information
    row = 3
    ws[f'A{row}'] = 'Student Information'
    ws[f'A{row}'].font = bold_font
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    student_info = [
        ('Student Name:', f"{student_user.first_name} {student_user.last_name}"),
        ('Student ID:', student_user.id_number or 'N/A'),
        ('Program:', student_program),
        ('Year Level:', f"Year {profile.year_level}" if profile.year_level else 'N/A'),
        ('Generated On:', datetime.now().strftime('%B %d, %Y at %I:%M %p')),
    ]
    
    for label, value in student_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = value
        row += 1
    
    # Progress Summary
    row += 2
    ws[f'A{row}'] = 'Progress Summary'
    ws[f'A{row}'].font = bold_font
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    ws[f'A{row}'].alignment = center_align
    ws[f'B{row}'].alignment = center_align
    
    row += 1
    summary_data = [
        ('Completion Percentage', f"{completion_percentage:.1f}%"),
        ('Credits Completed', f"{completed_units} / {total_units}"),
        ('Credits Remaining', str(credits_remaining)),
        ('Current GPA', f"{current_gpa:.2f}"),
        ('Expected Graduation', profile.expected_graduation.strftime('%B %Y') if profile.expected_graduation else 'TBD'),
    ]
    
    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    # Helper function to add course table
    def add_course_table(start_row, title, courses, fill_color):
        current_row = start_row
        ws[f'A{current_row}'] = title
        ws[f'A{current_row}'].font = bold_font
        ws.merge_cells(f'A{current_row}:E{current_row}')
        
        current_row += 1
        headers = ['Course Code', 'Course Title', 'Units', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        current_row += 1
        for course in courses:
            ws.cell(row=current_row, column=1, value=course['code']).border = thin_border
            ws.cell(row=current_row, column=2, value=course['title']).border = thin_border
            ws.cell(row=current_row, column=3, value=course['units']).border = thin_border
            ws.cell(row=current_row, column=3).alignment = center_align
            ws.cell(row=current_row, column=4, value=course['status']).border = thin_border
            
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).fill = fill_color
            
            current_row += 1
        
        return current_row + 1
    
    # Add course tables
    row += 2
    if completed_courses:
        row = add_course_table(row, 'Completed Courses', completed_courses, success_fill)
    
    if in_progress_courses:
        row = add_course_table(row, 'Courses In Progress', in_progress_courses, warning_fill)
    
    if not_started_courses:
        row = add_course_table(row, 'Remaining Courses', not_started_courses, neutral_fill)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="degree_audit_{student_user.id_number}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    return response